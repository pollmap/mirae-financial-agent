#!/usr/bin/env python3
"""Profile the organizer-provided XLSX files without modifying them.

The script is intentionally read-only. It produces machine-readable audit files
that can be regenerated after every ingestion change.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import tempfile
import unicodedata
import zipfile
from collections import Counter, defaultdict
from collections.abc import Iterable
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DATE_8_RE = re.compile(r"^(19|20|99)\d{6}$")
NULL_SENTINELS = {"NULL", "NONE", "N/A", "NA", "NAN"}
UNIQUE_CAP = 10_000
EXPECTED_ZIP_SHA256 = "c3809aca73396f57242ded0188fa06a3d271bd4ad65010e53d5533efc7c18163"


DATASETS = {
    "PRBD01N001": {
        "group": "국내채권",
        "official_rows": 42_394,
        "official_columns": 40,
        "key_specs": [("PD_NO",)],
    },
    "PREF01N001": {
        "group": "국내 ETF",
        "official_rows": 1_734,
        "official_columns": 73,
        "key_specs": [
            ("pd_itm_no",),
            ("pd_itm_no_ma",),
            ("pd_exg_mkt_cd", "pd_itm_no", "pd_itm_no_ma"),
        ],
    },
    "PREF02N001": {
        "group": "해외 ETF",
        "official_rows": 5_646,
        "official_columns": 49,
        "key_specs": [("pd_itm_no",), ("pd_itm_no_ma",), ("pd_isin_cd",)],
    },
    "PRFD01N001": {
        "group": "공모펀드",
        "official_rows": 95_619,
        "official_columns": 45,
        "key_specs": [
            ("itm_no",),
            ("itm_no", "prfd_attr_cd"),
            ("std_itm_no",),
            ("ksd_itm_no",),
        ],
    },
}


def normalized_text(value: Any) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFC", str(value)).strip()


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def decimal_value(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        text = str(value).strip().replace(",", "")
        if not text or text.upper() in NULL_SENTINELS:
            return None
        result = Decimal(text)
        return result if result.is_finite() else None
    except (InvalidOperation, ValueError):
        return None


def serializable(value: Any) -> Any:
    if value is None or isinstance(value, str | int | float | bool):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return str(value)
        return value
    return str(value)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


@dataclass
class FieldStats:
    name: str
    total: int = 0
    blank: int = 0
    whitespace_only: int = 0
    sentinel_null: int = 0
    numeric_parseable: int = 0
    zero: int = 0
    negative: int = 0
    numeric_min: Decimal | None = None
    numeric_max: Decimal | None = None
    date_min: str | None = None
    date_max: str | None = None
    samples: list[Any] = field(default_factory=list)
    uniques: set[str] = field(default_factory=set)
    unique_capped: bool = False
    top_values: Counter[str] = field(default_factory=Counter)

    def add(self, value: Any) -> None:
        self.total += 1
        if value is None:
            self.blank += 1
            return
        if isinstance(value, str) and value.strip() == "":
            self.blank += 1
            self.whitespace_only += 1
            return

        text = normalized_text(value)
        if text.upper() in NULL_SENTINELS:
            self.sentinel_null += 1
        if len(self.samples) < 3 and text not in {str(v) for v in self.samples}:
            self.samples.append(serializable(value))

        if not self.unique_capped:
            self.uniques.add(text)
            if len(self.uniques) > UNIQUE_CAP:
                self.unique_capped = True
                self.uniques.clear()
        if len(self.top_values) <= 200 or text in self.top_values:
            self.top_values[text] += 1

        number = decimal_value(value)
        if number is not None:
            self.numeric_parseable += 1
            if number == 0:
                self.zero += 1
            if number < 0:
                self.negative += 1
            self.numeric_min = number if self.numeric_min is None else min(self.numeric_min, number)
            self.numeric_max = number if self.numeric_max is None else max(self.numeric_max, number)

        # Date-range inference is limited to date-labelled fields. Applying an
        # eight-digit heuristic to amounts or identifiers creates false dates.
        column_lower = self.name.lower()
        is_date_field = column_lower.endswith("_dt") or "date" in column_lower
        compact = text.replace("-", "").replace("/", "")[:8]
        if is_date_field and DATE_8_RE.fullmatch(compact):
            self.date_min = compact if self.date_min is None else min(self.date_min, compact)
            self.date_max = compact if self.date_max is None else max(self.date_max, compact)

    def record(self, dataset_id: str, schema: dict[str, Any]) -> dict[str, Any]:
        nonblank = self.total - self.blank
        unique = f">{UNIQUE_CAP}" if self.unique_capped else len(self.uniques)
        return {
            "dataset_id": dataset_id,
            "column_name": self.name,
            "schema_type": schema.get("schema_type"),
            "column_name_ko": schema.get("column_name_ko"),
            "schema_example": schema.get("schema_example"),
            "total_rows": self.total,
            "nonblank_count": nonblank,
            "coverage_ratio": round(nonblank / self.total, 6) if self.total else 0,
            "blank_count": self.blank,
            "whitespace_only_count": self.whitespace_only,
            "sentinel_null_count": self.sentinel_null,
            "numeric_parseable_count": self.numeric_parseable,
            "zero_count": self.zero,
            "negative_count": self.negative,
            "unique_count": unique,
            "numeric_min": str(self.numeric_min) if self.numeric_min is not None else None,
            "numeric_max": str(self.numeric_max) if self.numeric_max is not None else None,
            "date_like_min": self.date_min,
            "date_like_max": self.date_max,
            "sample_values": json.dumps(self.samples, ensure_ascii=False),
            "top_values": json.dumps(self.top_values.most_common(10), ensure_ascii=False),
        }


def dataset_id_from_name(path: Path) -> str:
    for dataset_id in DATASETS:
        if path.name.startswith(dataset_id):
            return dataset_id
    raise ValueError(f"Unknown dataset file: {path.name}")


def read_schema(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Sheet1_Schema"]
    result: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if not row or is_blank(row[0]):
            continue
        result[normalized_text(row[0])] = {
            "pk_fk": serializable(row[1] if len(row) > 1 else None),
            "schema_type": serializable(row[2] if len(row) > 2 else None),
            "column_name_ko": serializable(row[3] if len(row) > 3 else None),
            "schema_example": serializable(row[4] if len(row) > 4 else None),
        }
    workbook.close()
    return result


def key_string(row: tuple[Any, ...], indexes: tuple[int, ...]) -> str | None:
    parts = [normalized_text(row[index]) for index in indexes]
    return "\x1f".join(parts) if all(parts) else None


def profile_dataset(data_path: Path, schema_path: Path) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    dataset_id = dataset_id_from_name(data_path)
    config = DATASETS[dataset_id]
    schema = read_schema(schema_path)
    workbook = load_workbook(data_path, read_only=True, data_only=True)
    sheet = workbook["datarows"]
    iterator = sheet.iter_rows(values_only=True)
    headers = tuple(normalized_text(v) for v in next(iterator))
    header_index = {name: index for index, name in enumerate(headers)}
    stats = [FieldStats(name) for name in headers]
    key_sets: dict[tuple[str, ...], set[str]] = {spec: set() for spec in config["key_specs"]}
    key_missing: Counter[tuple[str, ...]] = Counter()
    row_hashes: set[str] = set()
    exact_duplicate_rows = 0
    row_count = 0

    domain_counts: dict[str, Counter[str]] = defaultdict(Counter)
    selected_categorical = {
        "PRBD01N001": ["PD_EXG_MKT", "STD_PD_MCLS_NM", "STD_PD_SCLS_NM", "CURR_CD", "PD_RISK_GCD"],
        "PREF01N001": ["pd_grp_no", "pd_sale_yn", "pd_tr_yn", "pd_risk_nm", "wu_inv_ast_type", "wu_inv_rgn"],
        "PREF02N001": ["pd_grp_no", "cu_etn_yn", "pd_curr_cd", "pd_exg_mkt_cd", "wu_inv_ast_type", "wu_inv_rgn"],
        "PRFD01N001": ["or_attr_desc", "prvo_pbff_desc", "sale_yn", "zrin_fd_ivst_risk_grd_nm", "fd_ivst_rgn_desc"],
    }[dataset_id]

    for row in iterator:
        row_count += 1
        padded = tuple(row) + (None,) * max(0, len(headers) - len(row))
        for index, field_stat in enumerate(stats):
            field_stat.add(padded[index])

        digest = hashlib.blake2b(
            json.dumps([serializable(v) for v in padded[: len(headers)]], ensure_ascii=False, separators=(",", ":")).encode("utf-8"),
            digest_size=16,
        ).hexdigest()
        if digest in row_hashes:
            exact_duplicate_rows += 1
        else:
            row_hashes.add(digest)

        for spec in config["key_specs"]:
            indexes = tuple(header_index[name] for name in spec)
            value = key_string(padded, indexes)
            if value is None:
                key_missing[spec] += 1
            else:
                key_sets[spec].add(value)

        for column in selected_categorical:
            value = normalized_text(padded[header_index[column]])
            domain_counts[column][value or "<BLANK>"] += 1

    workbook.close()

    key_records = []
    for spec in config["key_specs"]:
        distinct = len(key_sets[spec])
        missing = key_missing[spec]
        key_records.append(
            {
                "dataset_id": dataset_id,
                "key_columns": "+".join(spec),
                "row_count": row_count,
                "nonblank_key_count": row_count - missing,
                "missing_key_count": missing,
                "distinct_key_count": distinct,
                "duplicate_key_rows": row_count - missing - distinct,
                "is_unique_when_present": row_count - missing == distinct,
            }
        )

    field_records = [stat.record(dataset_id, schema.get(stat.name, {})) for stat in stats]
    dataset_record = {
        "dataset_id": dataset_id,
        "product_group": config["group"],
        "source_file": data_path.name,
        "row_count": row_count,
        "column_count": len(headers),
        "official_row_count": config["official_rows"],
        "official_count_matches": row_count == config["official_rows"],
        "exact_duplicate_rows": exact_duplicate_rows,
        "schema_columns": len(schema),
        "header_schema_set_match": set(headers) == set(schema),
    }
    domain = {
        "dataset_id": dataset_id,
        "categorical_counts": {
            column: counter.most_common() for column, counter in domain_counts.items()
        },
    }
    return dataset_record, field_records, key_records, domain


def write_csv(path: Path, records: Iterable[dict[str, Any]]) -> None:
    records = list(records)
    if not records:
        return
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(records[0]))
        writer.writeheader()
        writer.writerows(records)


def find_pair(root: Path, dataset_id: str) -> tuple[Path, Path]:
    data_candidates = list(root.rglob(f"{dataset_id}*datarows.xlsx"))
    schema_candidates = list(root.rglob(f"{dataset_id}*schema.xlsx"))
    if len(data_candidates) != 1 or len(schema_candidates) != 1:
        raise RuntimeError(
            f"Expected exactly one datarows/schema pair for {dataset_id}; "
            f"got data={len(data_candidates)}, schema={len(schema_candidates)}"
        )
    return data_candidates[0], schema_candidates[0]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--zip", type=Path, required=True, help="Organizer-provided ZIP")
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--expected-zip-sha256",
        default=EXPECTED_ZIP_SHA256,
        help="Fail closed unless the organizer ZIP matches this SHA-256",
    )
    args = parser.parse_args()
    if not args.zip.is_file():
        raise FileNotFoundError(args.zip)
    args.output.mkdir(parents=True, exist_ok=True)

    source_sha256 = sha256_file(args.zip)
    if source_sha256 != args.expected_zip_sha256:
        raise RuntimeError(
            f"ZIP SHA-256 mismatch: expected {args.expected_zip_sha256}, got {source_sha256}"
        )
    with tempfile.TemporaryDirectory(prefix="mirae-profile-") as temp_dir:
        source_root = Path(temp_dir)
        with zipfile.ZipFile(args.zip) as archive:
            xlsx_members = [name for name in archive.namelist() if name.lower().endswith(".xlsx")]
            if len(xlsx_members) != 8:
                raise RuntimeError(f"Expected exactly 8 XLSX members, got {len(xlsx_members)}")
            for member in archive.infolist():
                resolved = (source_root / member.filename).resolve()
                if source_root.resolve() not in resolved.parents and resolved != source_root.resolve():
                    raise RuntimeError(f"Unsafe ZIP member path: {member.filename}")
            archive.extractall(source_root)

        dataset_records: list[dict[str, Any]] = []
        field_records: list[dict[str, Any]] = []
        key_records: list[dict[str, Any]] = []
        domain_records: list[dict[str, Any]] = []
        for dataset_id in DATASETS:
            data_path, schema_path = find_pair(source_root, dataset_id)
            dataset, fields, keys, domain = profile_dataset(data_path, schema_path)
            dataset_records.append(dataset)
            field_records.extend(fields)
            key_records.extend(keys)
            domain_records.append(domain)

        failures = []
        for record in dataset_records:
            config = DATASETS[record["dataset_id"]]
            if not record["official_count_matches"]:
                failures.append(f"{record['dataset_id']} row count")
            if record["column_count"] != config["official_columns"]:
                failures.append(f"{record['dataset_id']} column count")
            if not record["header_schema_set_match"]:
                failures.append(f"{record['dataset_id']} schema/header set")
        if sum(item["row_count"] for item in dataset_records) != 145_393:
            failures.append("total row count")
        if failures:
            raise RuntimeError("Source invariant failure: " + ", ".join(failures))

    write_csv(args.output / "dataset_profile.csv", dataset_records)
    write_csv(args.output / "field_profile.csv", field_records)
    write_csv(args.output / "key_profile.csv", key_records)
    summary = {
        "source_zip": args.zip.name,
        "source_sha256": source_sha256,
        "total_rows": sum(item["row_count"] for item in dataset_records),
        "datasets": dataset_records,
        "keys": key_records,
        "domain_counts": domain_records,
    }
    (args.output / "profile_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps({"status": "ok", "output": str(args.output), "total_rows": summary["total_rows"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
